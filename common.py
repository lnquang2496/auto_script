from openpyxl import load_workbook
from os       import rename, remove
from sys      import exit
from os.path  import isfile, join

def coor_find_cell(ws, target, match_case:bool=False, find_all:bool=False):
	return_value = list()
	target_found = False
	# If target is string type
	if isinstance(target, str):
		for rows in ws:
			for cell in rows:
				if match_case:
					target_found = target in cell.value
				else:
					target_found = target == cell.value

				if target_found:
					not_merged_cell = True
					for merged_cell in ws.merged_cells:
						if (cell.coordinate in merged_cell):
							temp = {'firstcol':merged_cell.bounds[0],
									'firstrow':merged_cell.bounds[1],
									'lastcol':merged_cell.bounds[2],
									'lastrow':merged_cell.bounds[3]}
							not_merged_cell = False
							if find_all:
								return_value.append(temp)
							else:
								return dict(temp)
					if not_merged_cell:
						temp = {'firstcol':cell.column,
								'firstrow':cell.row,
								'lastcol':cell.column,
								'lastrow':cell.row}
						if find_all:
							return_value.append(temp)
						else:
							return dict(temp)
	# If target is list type, list format [column, row]
	elif isinstance(target, list):
		# list  = [col, row]
		coor = ws.cell(column=target[0], row=target[1]).coordinate
		not_merged_cell = True
		for merged_cell in ws.merged_cells:
			if (coor in merged_cell):
				temp = {'firstcol':merged_cell.bounds[0],
						'firstrow':merged_cell.bounds[1],
						'lastcol':merged_cell.bounds[2],
						'lastrow':merged_cell.bounds[3]}
				not_merged_cell = False
				return dict(temp)
		if not_merged_cell:
			temp = {'firstcol':target[0],
					'firstrow':target[1],
					'lastcol':target[0],
					'lastrow':target[1]}
			return dict(temp)
	return return_value

def get_cell_value(ws, coor)->str:
	if isinstance(coor, list):
		val = ws.cell(column=coor[0], row=coor[1]).value
		return str(val)
	else:
		for row in range (coor['firstrow'], coor['lastrow'] + 1):
			for col in range (coor['firstcol'], coor['lastcol'] + 1):
				val = ws.cell(column=col, row=row).value
				if (val != None):
					return str(val)
	return None

def coor_shift_right(ws, coor):
	temp = list([coor['lastcol'] + 1, coor['firstrow']])
	return coor_find_cell(ws, temp)

def coor_shift_left(ws, coor):
	col = coor['firstcol'] - 1
	if col < 1:
		col = 1
	temp = list([col, coor['firstrow']])
	return coor_find_cell(ws, temp)

def coor_shift_up(ws, coor):
	row = coor['firstrow'] - 1
	if row < 1:
		row = 1
	temp = list([coor['firstcol'], row])
	return coor_find_cell(ws, temp)

def coor_shift_down(ws, coor):
	temp = list([coor['firstcol'], coor['lastrow'] + 1])
	return coor_find_cell(ws, temp)

def load_worksheet(filename:str, sheetname:str):
	try:
		wb = load_workbook(filename, data_only=True)
	except:
		exit(0)
	ws = wb[sheetname]
	return ws

def count_cell_in_merged_cell(ws, cell_1)->int:
	count = 0
	cell_2 = coor_shift_down(ws, cell_1)
	while cell_2['lastcol'] <= cell_1['lastcol']:
		count += 1
		cell_2 = coor_shift_right(ws, cell_2)
	return count

def get_type_name(ws, value:str)->str:
	temp_val = value.replace('[a]', '').replace('[g]', '').\
		replace('*', '').replace(';', '').replace('  ', ' ')
	all_val = temp_val.split(' ')
	cell_name = all_val[len(all_val) - 1]
	cell_type = ''
	for i in range(len(all_val) - 1):
		cell_type += f'{all_val[i]} '
	return cell_type[:-1], cell_name

def check_pointer(ws, value:str)->bool:
	valid_pointer = ['*', 'IODevice ', 'DevTree_Node']
	for x in valid_pointer:
		if x in value:
			return True
	return False

def check_structure(ws, value:str)->bool:
	valid_structure= ['struct', 'imr_state_mng', 'imr_rtt_state_mng', 'imr_struct_t']
	for x in valid_structure:
		if x in value:
			return True
	return False

def row_of_testcase(ws, symbol):
	cur_cell = coor_find_cell(ws, symbol)
	col_of_tc = cur_cell['firstcol']
	first_row_of_tc = cur_cell['lastrow'] + 1
	while (get_cell_value(ws, cur_cell) != None):
		cur_cell = coor_shift_down(ws, cur_cell)
	last_row_of_tc = cur_cell['firstrow'] - 1
	return first_row_of_tc, last_row_of_tc, col_of_tc

def select_check_type(ws, value:str)->str:
	check_address = ['IODevice ', 'char', 'Clock', 'Semaphore', 'Task']
	check_u_int = ['uint', 'unsigned', 'Address']
	check_bool = ['bool', 'Boolean']
	# default check type is CHECK_S_INT
	check_type = 'CHECK_S_INT'
	for c in check_u_int:
		if c in value:
			check_type = 'CHECK_U_INT'
	for c in check_bool:
		if c in value:
			check_type = 'CHECK_BOOLEAN'
	for c in check_address:
		if c in value:
			check_type = 'CHECK_ADDRESS'
	return check_type

def file_append(f_path:str, f_name:str, data:str, position, append_type:bool = False):
	file_targ = join(f_path, f_name)
	file_temp = join(f_path, f_name + "_temp.txt")
	file_old  = join(f_path, f_name + "_old.txt")
	pos_count = 0

	def append_data(line:str, data:str):
		return data + line if append_type else line + data

	if not isfile(file_targ):
		exit(f"File \"{file_targ}\" not exist")

	if isfile(file_temp):
		remove(file_temp)

	with open(file_targ, "r") as i, open(file_temp, "w") as o:
		for index, l in enumerate(i):
			backup_l = l
			if isinstance(position, int):
				if position == index:
					l = append_data(l, data)

			elif isinstance(position, str):
				if position in l:
					l = append_data(l, data)

			elif isinstance(position, list):
				if pos_count < len(position):
					if position[pos_count] in l:
						pos_count += 1
					if pos_count == len(position):
						l = append_data(l, data)
			try:
				o.write(l)
			except:
				o.write(backup_l)
				print(l)

	if isfile(file_old):
		remove(file_old)

	rename(file_targ, file_old)
	rename(file_temp, file_targ)

def file_clear(f_path:str, f_name:str, start_pos, end_pos):
	file_targ = join(f_path, f_name)
	file_temp = join(f_path, f_name + "_temp.txt")
	file_old  = join(f_path, f_name + "_old.txt")
	start_pos_count = 0
	end_pos_count = 0
	flag_clear = False

	if not isfile(file_targ):
		exit(f"File \"{file_targ}\" not exist")

	if isfile(file_temp):
		remove(file_temp)

	with open(file_targ, "r") as i, open(file_temp, "w") as o:
		for index, l in enumerate(i, 1):
			if isinstance(start_pos, int) and isinstance(end_pos, int):
				if (start_pos < index) and (end_pos > index):
					l = ""

			elif isinstance(start_pos, str) and isinstance(end_pos, str):
				if start_pos in l:
					flag_clear = True
					o.write(l)
				if end_pos in l:
					flag_clear = False

			elif isinstance(start_pos, list) and isinstance(end_pos, list):
				if start_pos_count < len(start_pos):
					if start_pos[start_pos_count] in l:
						start_pos_count += 1
					if start_pos_count == len(start_pos):
						flag_clear = True
						o.write(l)
				if flag_clear:
					if end_pos_count < len(end_pos):
						if end_pos[end_pos_count] in l:
							end_pos_count += 1
						if end_pos_count == len(end_pos):
							flag_clear = False
			
			if flag_clear:
				l = ""

			o.write(l)

	if isfile(file_old):
		remove(file_old)

	rename(file_targ, file_old)
	rename(file_temp, file_targ)

# Extract data from PCL with specific target
def get_data(ws, target, input_range, cur_row):
	data = ''
	input_cell = coor_shift_down(ws, input_range)
	while input_cell['lastcol'] <= input_range['lastcol']:
		# Check input cell merged range, if merge range use {var1, var2}
		if (target in get_cell_value(ws, input_cell)):
			if (input_cell['lastcol'] - input_cell['firstcol']) == 0:
				cur_input_param = get_cell_value(ws, [input_cell['firstcol'], cur_row])
				data += f'{cur_input_param}, '
			else:
				element_cell = coor_shift_down(ws, input_cell)
				while element_cell['lastcol'] <= input_cell['lastcol']:
					cur_input_param = get_cell_value(ws, [element_cell['firstcol'], cur_row])
					data += f'{cur_input_param}, '
					element_cell = coor_shift_right(ws, element_cell)

		input_cell = coor_shift_right(ws, input_cell)
	return data

# Extract data from PCL with specific target
def get_data_with_bralet(ws, target, input_range, cur_row):
	data = ''
	input_cell = coor_shift_down(ws, input_range)
	while input_cell['lastcol'] <= input_range['lastcol']:
		# Check input cell merged range, if merge range use {var1, var2}
		if (target in get_cell_value(ws, input_cell)):
			if (input_cell['lastcol'] - input_cell['firstcol']) == 0:
				cur_input_param = get_cell_value(ws, [input_cell['firstcol'], cur_row])
				data += f'{cur_input_param}, '
			else:

				data = f'{data}{{'
				element_cell = coor_shift_down(ws, input_cell)
				while element_cell['lastcol'] <= input_cell['lastcol']:
					cur_input_param = get_cell_value(ws, [element_cell['firstcol'], cur_row])
					data += f'{cur_input_param}, '
					element_cell = coor_shift_right(ws, element_cell)

				data = f'{data[:-2]}}}, '

		input_cell = coor_shift_right(ws, input_cell)
	return data

# Create file .h contrain test case
def create_test_case_file(ws, worksheet, src_dir, src, check_sequence):
	# Get Test case start row and end row
	start_row, end_row, testcase_col = row_of_testcase(ws, '#')
	# Get Input factor range
	input_factor = coor_find_cell(ws, 'Input factor')
	# Get Output element range
	output_element = coor_find_cell(ws, 'Output element')
	# Create file .h

	dot_h_dir = f'{src_dir}test_{worksheet}.h'
	dot_h = open(dot_h_dir, 'w')

	# Begin of file
	data = 'static struct CPPTH_LOOP_INPUT_STRUCT CPPTH_LOOP_INPUT[] = {\n'

	def get_func_name(mark:str, target:str)->str:
		func_full = target.replace(mark, '')
		func_type = func_full[ : func_full.find(' ')]
		func_name = func_full[ func_full.find(' ') + 1 : func_full.find('(')]
		return func_type, func_name

	dot_h.write(data)
	# Add test case
	# Input factor
	for cur_row in range(start_row, end_row + 1):
		# Reset data
		data = '\t{'
		# Add test case number to data
		tc_num = get_cell_value(ws, [testcase_col, cur_row])
		tc_num = tc_num[:tc_num.find('-')] + '_' + tc_num[tc_num.find('-') + 1:]
		data += f'\"{tc_num}\", '
		# Add description to data - Named: Item
		describe = f'{worksheet}_{tc_num}'
		data += f'\"{describe}\", '
		del describe

		# Add expected calls sequence
		# In case not check sequence of calling stub function
		# TODO: Add feature: many instance in sequence

		data += '"'
		CELL = coor_shift_down(ws, input_factor)
		list_of_function_called = list()
		count = 0
		while CELL['lastcol'] <= input_factor['lastcol']:
			CELL_VAL = get_cell_value(ws, CELL)
			if ('[rt]' in CELL_VAL):
				# Get function name from [rt]Error MyFunction(int a, int b);
				#                                  MyFunction
				FTYPE, FNAME = get_func_name('[rt]', CELL_VAL)
				'''
				FNAME = CELL_VAL[CELL_VAL.find(' ') + 1 : CELL_VAL.find('(')]
				del CELL_VAL
				# Check loop of instance
				if (FNAME not in dict(list_of_function_called)):
					list_of_function_called.append([FNAME, 0])
				else:
					function_count = list_of_function_called[list(dict(list_of_function_called)).index(FNAME)][1]
					list_of_function_called[list(dict(list_of_function_called)).index(FNAME)][1] = function_count + 1
				'''
				# Check is this instance called
				FRETVAL = get_cell_value(ws, [CELL['firstcol'], cur_row])

				FUNCINSTANCE = ''
				if (FRETVAL != None) and (FRETVAL != '-'):
					FUNCINSTANCE = f'{FNAME}#{tc_num}_{count}'
					if check_sequence == True:
						data += f'{FUNCINSTANCE}; '
					else:
						data += f'{{{FUNCINSTANCE}}} '
				count += 1
			CELL = coor_shift_right(ws, CELL)
		data = f'{data}\", '

		# Add execute - 1: execute this function
		judgment = coor_find_cell(ws, 'Judgment')
		if ('Exclude' in get_cell_value(ws, [judgment['firstcol'], cur_row])):
			data = f'{data}0, '
		else:
			data = f'{data}1, '

		# Add input param by detect [a]
		# TODO: Add detect structure - DONE

		data += get_data(ws, '[a]', input_factor, cur_row)

		# Add global variable by detect [g]
		data += get_data(ws, '[g]', input_factor, cur_row)

		if ("FillRect" in worksheet):
			temp_data = get_data_with_bralet(ws, '[a]', output_element, cur_row)
		else:
			temp_data = get_data(ws, '[a]', output_element, cur_row)
		# TODO: Add detect check [a] param output - DONE
		#temp_data = get_data(ws, '[a]', output_element, cur_row)
		if '-,' in temp_data:
			temp_data = temp_data.replace('-,', 'DONTCARE,')
		data += temp_data
		del temp_data

		# Add expected global variable by detect [g] in output element
		data += get_data(ws, '[g]', output_element, cur_row)

		# Add test result by detect 'Return value' in output element
		data += get_data(ws, 'Return value', output_element, cur_row)

		# Write all the data to file
		data = f'{data[:-2]}}},\n'
		dot_h.write(data)

	# End of file
	data = '};\n'
	dot_h.write(data)
	dot_h.close()
	del dot_h

# Create stub instance
def create_stub_file(ws, worksheet:str, src_dir:str, src:str):
	start_row, end_row, testcase_col = row_of_testcase(ws, '#')
	cell_1_left                      = coor_find_cell(ws, 'Input factor')
	cell_1_right                     = coor_find_cell(ws, 'Output element')

	def get_func_name(mark:str, target:str)->str:
		func_full = target.replace(mark, '')
		func_type = func_full[ : func_full.find(' ')]
		func_name = func_full[ func_full.find(' ') + 1 : func_full.find('(')]
		return func_type, func_name

	def get_check_type(val:str)->str:
		type_1 = ['true', 'false']

		if  ("_CH") in val:
			return "CHECK_ADDRESS"

		if ("OFFSET") in val:
			return "CHECK_U_INT"

		for i in type_1:
			if i in val:
				return 'CHECK_BOOLEAN'

		if '0x' in val:
			if val.endswith('u') or val.endswith('U'):
				return 'CHECK_U_INT'
			else:
				return 'CHECK_S_INT'

		if val.endswith('u') or val.endswith('U'):
			if val.replace('u', '').replace('U', '').isdigit:
				return 'CHECK_U_INT'

		try:
			a = int(val)
		except:
			pass
		else:
			return 'CHECK_S_INT'

		if val.upper:
			return 'CHECK_ADDRESS'

		return 'CHECK_S_INT'

	for cur_row in range(start_row, end_row + 1):
		func_list = list()
		tc_num = get_cell_value(ws, [testcase_col, cur_row]).replace('-', '_')
		count = 0

		cell_2_left = coor_shift_down(ws, cell_1_left)
		while (cell_2_left['lastcol'] <= cell_1_left['lastcol']):
			cell_2_left_val = get_cell_value(ws, cell_2_left)

			if '[rt]' in cell_2_left_val:
				func_type, func_name = get_func_name('[rt]', cell_2_left_val)
				cur_val = get_cell_value(ws, [cell_2_left['firstcol'], cur_row])

				title = f'/* Isolate for function {func_name} */\n'

				func_list.append(f'{func_name}_{count}')

				# Set instance_data
				instance_data = f'\tIF_INSTANCE(\"{tc_num}_{count}\") {{\n'

				# START: Set return value of funtion
				if cur_val == 'None' or cur_val == '-':
					pass
				else:
					if 'void' in func_type:
						return_data = '\t\treturn;\n'
					else:
						return_data = f'\t\treturn {cur_val};\n'
					# END: Set return value of funtion
					# START: Set output value of function
					temp_cell_2_left = coor_shift_right(ws, cell_2_left)
					temp_cell_2_val  = get_cell_value(ws, temp_cell_2_left)
					output_data = str()
					if '[f]' in temp_cell_2_val and func_name in temp_cell_2_val and temp_cell_2_left['lastcol'] <= cell_1_left['lastcol']:
						cell_2_left = temp_cell_2_left
						cell_3_left = coor_shift_down(ws, cell_2_left)
						while cell_3_left['lastcol'] <= cell_2_left['lastcol']:
							cell_3_left_val = get_cell_value(ws, cell_3_left)
							cur_val = get_cell_value(ws, [cell_3_left['firstcol'], cur_row])

							if cur_val != 'None' and cur_val != '-':
								if 'UTS' in cur_val:
									if cur_val.find(')') != -1:
										cast_type = cur_val[:cur_val.find(')') + 1]
									else:
										cast_type = ''

									if '*' in cell_3_left_val:
										output_data = f'{output_data}\t\t{cell_3_left_val} = {cast_type}&local;\n'
									else:
										output_data = f'{output_data}\t\t*{cell_3_left_val} = {cast_type}&local;\n'
								else:
									if '->' in cell_3_left_val:
										output_data = f'{output_data}\t\t{cell_3_left_val} = {cur_val};\n'
									else:
										if '*' in cell_3_left_val:
											output_data = f'{output_data}\t\t{cell_3_left_val} = {cur_val};\n'
										else:
											output_data = f'{output_data}\t\t*{cell_3_left_val} = {cur_val};\n'

							cell_3_left = coor_shift_right(ws, cell_3_left)
					# END: Set output value of function
					# START: Output element process
					temp_count = 0
					temp_func = str()
					check_data = str()
					cell_2_right = coor_shift_down(ws, cell_1_right)
					while (cell_2_right['lastcol'] <= cell_1_right['lastcol']):
						cell_2_right_val = get_cell_value(ws, cell_2_right)
						######
						if '[f]' in cell_2_right_val:
							temp_func_type, temp_func_name = get_func_name('[f]', cell_2_right_val)
							cell_3_right = coor_shift_down(ws, cell_2_right)

							temp_func = f'{temp_func_name}_{temp_count}'
							temp_count += 1

							process = False
							if func_list[count] == temp_func:
									process = True
							if process:
								while cell_3_right['lastcol'] <= cell_2_right['lastcol']:
									cell_3_right_val = get_cell_value(ws, cell_3_right)
									cur_val = get_cell_value(ws, [cell_3_right['firstcol'], cur_row])
									if cur_val != 'None' and cur_val != '-':
										if 'UTS_' in cur_val:
											if ('kbase' == cell_3_right_val) or ('vect' == cell_3_right_val):
												check_data = f'{check_data}\t\tCHECK_BOOLEAN(({cell_3_right_val} != 0), true);\n'
											else:
												check_data = f'{check_data}\t\tCHECK_BOOLEAN(({cell_3_right_val} != NULL), true);\n'
										elif 'NULL' in cur_val:
											check_data = f'{check_data}\t\tCHECK_BOOLEAN(({cell_3_right_val} == NULL), true);\n'
										else:
											check_type = get_check_type(cur_val)
											check_data = f'{check_data}\t\t{check_type}({cell_3_right_val}, {cur_val});\n'

									cell_3_right = coor_shift_right(ws, cell_3_right)
						cell_2_right = coor_shift_right(ws, cell_2_right)
					# END: Output element process

					final_data = f'{instance_data}{check_data}{output_data}{return_data}\t}}\n'
					del instance_data, check_data, output_data, return_data

					position = [title, 'IF_INSTANCE("default")', '}', 'LOG_SCRIPT_ERROR']

					file_append(src_dir, f"test_{src}.c", final_data, position, True)

				count += 1
			cell_2_left = coor_shift_right(ws, cell_2_left)

def pcl_to_testprogram(ws):
	list_of_input = []
	data_1 = ''
	data_2 = ''
	data_3 = ''
	data_4 = ''

	cell_1 = coor_find_cell(ws, 'Input factor')
	cell_2 = coor_shift_down(ws, cell_1)
	while cell_2['lastcol'] <= cell_1['lastcol']:
		cell_2_val = get_cell_value(ws, cell_2)

		if '[g]' in cell_2_val:
			is_global = True
		else:
			is_global = False

		if '[a]' in cell_2_val:
			is_argument = True
		else:
			is_argument = False

		if is_global == True or is_argument == True:
			cell_2_val = cell_2_val.replace('[a]', '').replace('[g]', '')
			cell_2_type, cell_2_name = get_type_name(ws, cell_2_val)
			is_cell_2_pointer = check_pointer(ws, cell_2_val)
			is_cell_2_structure = check_structure(ws, cell_2_val)
			exist = False

			if cell_2_name not in list_of_input:
				list_of_input.append(cell_2_name)
			else:
				exist = True
			###
			if exist == False and is_global == False:
				temp_cell_2_val = cell_2_val.replace('[a]', '').replace('[g]', '').replace(';', '')
				temp_data = f'\t\t{temp_cell_2_val};\n'
				data_1 += f'{temp_data}'
				del temp_cell_2_val

			if (cell_2['lastcol'] - cell_2['firstcol'] == 0) and (exist == False):
				'''NOT MERGED CELL'''

				if is_cell_2_pointer:
					###
					temp_data = f'\t{cell_2_type} local_{cell_2_name};\n'
					data_2 += f'{temp_data}'
					###
					temp_data = f'\t\t\tif (CURRENT_TEST.{cell_2_name} != NULL){{\n\t\t\t\tCURRENT_TEST.{cell_2_name} = &local_{cell_2_name};\n\t\t\t}}\n'
					data_3 += f'{temp_data}'
				pass

			else:
				'''IS MERGED CELL'''
				cell_2_number = ''
				if '[' in cell_2_val:
					temp_cell_2_number = cell_2_val[cell_2_val.find('[') + 1 : cell_2_val.find(']')]
					cell_2_number = f'_{temp_cell_2_number}'
					cell_2_val = cell_2_val.replace('[', '_').replace(']', '')
					del temp_cell_2_number

				cell_3 = coor_shift_down(ws, cell_2)
				while cell_3['lastcol'] <= cell_2['lastcol']:
					cell_3_val = get_cell_value(ws, cell_3)
					cell_3_type, cell_3_name = get_type_name(ws, cell_3_val)
					is_cell_3_pointer = check_pointer(ws, cell_3_val)
					is_cell_3_structure = check_structure(ws, cell_3_val)
					###
					cell_3_name_new = cell_3_name
					if '[' in cell_3_name:
						number = cell_3_name[cell_3_name.find('[') + 1 : cell_3_name.find(']')]
						try:
							number = int(number)
							cell_2_number = f'_{number}'
							cell_3_name_new = cell_3_name[: cell_3_name.find('[')]

						except:
							start_temp = cell_3_name.find('[')
							cell_2_number += f'{cell_3_name[start_temp:]}'
							cell_3_name_new = cell_3_name[: cell_3_name.find('[')]
							del start_temp
					###
					if '[' in cell_3_val:
						cell_3_val = cell_3_val[: cell_3_val.find('[')]
					temp_data = f'\t\t{cell_3_val}{cell_2_number};\n'
					data_1 += f'{temp_data}'

					if is_cell_3_pointer:
						###
						temp_data = f'\t{cell_3_type} local_{cell_3_name_new};\n'
						data_2 += f'{temp_data}'
						###
						temp_data = '\t\t\tif (CURRENT_TEST.{name}{number} != NULL){{\n\t\t\t\tCURRENT_TEST.{name}{number} = &local_{name};\n\t\t\t}}\n'.format(name = cell_3_name_new, number = cell_2_number)
						data_3 += f'{temp_data}'
					###
					if is_global:
						# if is_cell_3_pointer:
						# 	access = '->'
						# else:
						# 	access = '.'
						access = '.'
						if ('char' in cell_3_type):
							temp_data = '\t\t\tstrcpy({}.{}, CURRENT_TEST.{}{});\n'.format(cell_2_name, cell_3_name, cell_3_name_new, cell_2_number)
						else:
							temp_data = '\t\t\t{}{}{} = CURRENT_TEST.{}{};\n'.format(cell_2_name, access, cell_3_name, cell_3_name_new, cell_2_number)
						data_3 += f'{temp_data}'

					if exist:
						temp_data = '\t\t\tlocal_{cell_2_name}.{cell_3_name} = CURRENT_TEST.{cell_3_name};\n'.format(\
							cell_2_name = cell_2_name, cell_3_name = cell_3_name_new)
						data_3 += f'{temp_data}'

					cell_3 = coor_shift_right(ws, cell_3)
				pass
		cell_2 = coor_shift_right(ws, cell_2)

	del list_of_input

	cell_1 = coor_find_cell(ws, 'Output element')
	cell_2 = coor_shift_down(ws, cell_1)
	while cell_2['lastcol'] <= cell_1['lastcol']:
		cell_2_val = get_cell_value(ws, cell_2)

		if '[g]' in cell_2_val:
			is_global = True
			cell_2_val = cell_2_val.replace('[g]', '')
		else:
			is_global = False

		if '[a]' in cell_2_val:
			is_argument = True
			cell_2_val = cell_2_val.replace('[a]', '')
		else:
			is_argument = False

		if is_global == True or is_argument == True:
			cell_2_type, cell_2_name = get_type_name(ws, cell_2_val)
			is_cell_2_pointer = check_pointer(ws, cell_2_val)
			is_cell_2_structure = check_structure(ws, cell_2_val)

			if cell_2['lastcol'] - cell_2['firstcol'] == 0:

				if is_cell_2_pointer:
					###
					temp_data = '\t\t{} expected_{};\n'.format(cell_2_type, cell_2_name)
					data_1 += f'{temp_data}'

					init_val = get_cell_value(ws, coor_shift_down(ws, cell_2))
					if init_val != '-' and init_val != None:
						###
						temp_data = '\t\t\tlocal_{} = {};\n'.format(cell_2_name, init_val)
						data_3 += f'{temp_data}'

					###
					check_type = select_check_type(ws, cell_2_val)
					'''
					temp_data = '\t\t\t{check}({left}, {right});\n'.format(\
						check = check_type,\
						left = 'local_{}'.format(cell_2_name),\
						right = 'CURRENT_TEST.expected_{}'.format(cell_2_name)\
					)
					'''
					temp_data =  f'\t\t\tif (CURRENT_TEST.expected_{cell_2_name} != DONTCARE) {{\n'
					temp_data += f'\t\t\t\t{check_type}(local_{cell_2_name}, CURRENT_TEST.expected_{cell_2_name});\n'
					temp_data += f'\t\t\t}}'
					data_4 += f'{temp_data}'

				if is_global:
					if '[' in cell_2_name:
						temp_data = '\t\t{} expected_{};\n'.format(cell_2_type, cell_2_name.replace("[", "_").replace("]", ""))
					else:
						temp_data = '\t\t{} expected_{};\n'.format(cell_2_type, cell_2_name)
					data_1 += f'{temp_data}'

					check_type = select_check_type(ws, cell_2_val)
					'''
					temp_data = '\t\t\t{check}({left}, {right});\n'.format(\
						check = check_type,\
						left = '{}'.format(cell_2_name),\
						right = 'CURRENT_TEST.expected_{}'.format(cell_2_name.replace("[", "_").replace("]", ""))\
					)
					'''
					temp_data = f'\t\t\t{check_type}({cell_2_name}, CURRENT_TEST.expected_{cell_2_name.replace("[", "_").replace("]", "")});\n'
					data_4 += f'{temp_data}'
					pass

				pass
			else:
				'''IS MERGED CELL'''

				cell_3 = coor_shift_down(ws, cell_2)
				while cell_3['lastcol'] <= cell_2['lastcol']:
					cell_3_val = get_cell_value(ws, cell_3)
					cell_3_type, cell_3_name = get_type_name(ws, cell_3_val)
					is_cell_3_pointer = check_pointer(ws, cell_3_val)
					is_cell_3_structure = check_structure(ws, cell_3_val)

					cell_2_number = ''
					if '[' in cell_2_val:
						cell_2_number = '_{}'.format(cell_2_val[cell_2_val.find('[') + 1 : cell_2_val.find(']')])
					###
					cell_3_name_new = cell_3_name
					if '[' in cell_3_name:
						number = cell_3_name[cell_3_name.find('[') + 1 : cell_3_name.find(']')]
						try:
							number = int(number)
							#cell_2_number = '{}{}'.format(cell_2_number, cell_3_name[cell_3_name.find('['):])
							cell_2_number = '_{}'.format(number)
							cell_3_name_new = cell_3_name[: cell_3_name.find('[')]

						except:
							cell_2_number = '{}{}'.format(cell_2_number, cell_3_name[cell_3_name.find('['):])
							cell_3_name_new = cell_3_name[: cell_3_name.find('[')]

					temp_data = '\t\t{} expected_{}{};\n'.format(cell_3_type, cell_3_name_new, cell_2_number)
					data_1 += f'{temp_data}'

					###
					if is_global:

						if is_cell_3_pointer:
							access = '->'
						else:
							access = '.'
						check_type = select_check_type(ws, cell_3_val)
						temp_data = '\t\t\t{check}({left}, {right});\n'.format(\
							check = check_type,\
							left = '{}{}{}'.format(cell_2_name, access, cell_3_name),\
							right = 'CURRENT_TEST.expected_{}{}'.format(cell_3_name_new, cell_2_number)\
						)
						data_4 += f'{temp_data}'

					cell_3 = coor_shift_right(ws, cell_3)
				pass

		cell_2 = coor_shift_right(ws, cell_2)

	return data_1[:-1], data_2, data_3, data_4

def get_input_argument(ws, top_cell_name:str, target:str)->str:
	valid_top_cell_name = ['Input factor', 'Output element']
	valid_target = ['[a]', '[f]', '[rt]', '[g]']
	if (target not in valid_target) or (top_cell_name not in valid_top_cell_name):
		return
	del valid_top_cell_name, valid_target

	cell_1 = coor_find_cell(ws, top_cell_name)
	cell_2 = coor_shift_down(ws, cell_1)
	data = ''
	list_of_handle = []
	while (cell_2['lastcol'] <= cell_1['lastcol']):
		### Begin of while
		cell_2_val = get_cell_value(ws, cell_2)
		if target in cell_2_val:
			cell_2_type, cell_2_name = get_type_name(ws, cell_2_val)
			is_cell_2_pointer = check_pointer(ws, cell_2_val)
			is_cell_2_structure = check_structure(ws, cell_2_val)
			if cell_2_name in list_of_handle:
				pass
			else:
				list_of_handle.append(cell_2_name)
				data_append = f'CURRENT_TEST.{cell_2_name},'
				data += f'{data_append}'

		###	Next of while
		cell_2 = coor_shift_right(ws, cell_2)
	return data[:-1]
	pass

def create_test_program(ws, func_name:str)->str:
	data_1, data_2, data_3, data_4 = pcl_to_testprogram(ws)
	input_argument = get_input_argument(ws, 'Input factor', '[a]')
	data = f'''\
void test_{func_name}(){{
	struct CPPTH_LOOP_INPUT_STRUCT {{
		/* Test case data declarations */
		char* name;
		char* description;
		char* expected_calls;
		int execute;
{data_1}
		int32_t expected_returnValue;
	}};
	int32_t returnValue;
	/* Import external data declarations */
	#include "test_{func_name}.h"
{data_2}
	START_TEST_LOOP();
		/* Expected Call Sequence  */
		EXPECTED_CALLS(CURRENT_TEST.expected_calls);
			/* Set global data */
			initialise_global_data();
			/* Set expected values for global data checks */
			initialise_expected_global_data();
{data_3}
			/* Call SUT */
			returnValue = {func_name}({input_argument});

			/* Test case checks */
			CHECK_S_INT(returnValue, CURRENT_TEST.expected_returnValue);
{data_4}
		END_CALLS();
	END_TEST_LOOP();
}}
'''
	return data
