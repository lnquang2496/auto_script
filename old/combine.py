import os, sys, getopt
# openpyxl = 2.6.2
from openpyxl import load_workbook
from shutil import copyfile

#------------------------------------------------------------------------------#
# Find cell boundary
def find_cell(ws, target, match_case=False, find_all=False):
	return_value = list()
	target_found = False
	# If target is string type
	if isinstance(target, str):
		for rows in ws:
			for cell in rows:
				if match_case:
					target_found = target in cell.value
				else:
					target_found = target == cell.value

				if target_found:
					not_merged_cell = True
					for merged_cell in ws.merged_cells:
						if (cell.coordinate in merged_cell):
							temp = {'firstcol':merged_cell.bounds[0],
									'firstrow':merged_cell.bounds[1],
									'lastcol':merged_cell.bounds[2],
									'lastrow':merged_cell.bounds[3]}
							not_merged_cell = False
							if find_all:
								return_value.append(temp)
							else:
								return dict(temp)
					if not_merged_cell:
						temp = {'firstcol':cell.column,
								'firstrow':cell.row,
								'lastcol':cell.column,
								'lastrow':cell.row}
						if find_all:
							return_value.append(temp)
						else:
							return dict(temp)
	# If target is list type, list format [column, row]
	elif isinstance(target, list):
		# list  = [col, row]
		coor = ws.cell(column=target[0], row=target[1]).coordinate
		not_merged_cell = True
		for merged_cell in ws.merged_cells:
			if (coor in merged_cell):
				temp = {'firstcol':merged_cell.bounds[0],
						'firstrow':merged_cell.bounds[1],
						'lastcol':merged_cell.bounds[2],
						'lastrow':merged_cell.bounds[3]}
				not_merged_cell = False
				return dict(temp)
		if not_merged_cell:
			temp = {'firstcol':target[0],
					'firstrow':target[1],
					'lastcol':target[0],
					'lastrow':target[1]}
			return dict(temp)

	return return_value
#------------------------------------------------------------------------------#
# Get cell value with cell coordinate
def get_cell_value(ws, coor)->str:
	if isinstance(coor, list):
		val = ws.cell(column=coor[0], row=coor[1]).value
		return val
	else:
		for row in range (coor['firstrow'], coor['lastrow'] + 1):
			for col in range (coor['firstcol'], coor['lastcol'] + 1):
				val = ws.cell(column=col, row=row).value
				if (val != None):
					return val
	return None
#------------------------------------------------------------------------------#
# Shift 'right' cell
def coor_shift_right(ws, coor):
	temp = list([coor['lastcol'] + 1, coor['firstrow']])
	return find_cell(ws, temp)
#------------------------------------------------------------------------------#
# Shift 'left' cell
def coor_shift_left(ws, coor):
	col = coor['firstcol'] - 1
	if col < 1:
		col = 1
	temp = list([col, coor['firstrow']])
	return find_cell(ws, temp)
#------------------------------------------------------------------------------#
# Shift 'up' cell
def coor_shift_up(ws, coor):
	row = coor['firstrow'] - 1
	if row < 1:
		row = 1
	temp = list([coor['firstcol'], row])
	return find_cell(ws, temp)
#------------------------------------------------------------------------------#
# Shift 'down' cell
def coor_shift_down(ws, coor):
	temp = list([coor['firstcol'], coor['lastrow'] + 1])
	return find_cell(ws, temp)
#------------------------------------------------------------------------------#
def load_worksheet(filename, sheetname):
	try:
		wb = load_workbook(filename, data_only=True)
	except:
		sys.exit(0)
	ws = wb[sheetname]
	return ws
#------------------------------------------------------------------------------#
def get_type_name(ws, target:str):
	valid_pointer = ['*', 'IODevice']
	valid_structure = ['struct', 'g_struct', 'imr_state_mng', 'imr_rtt_state_mng']
	is_pointer = False
	is_structure = False
	for x in valid_pointer:
		if x in target:
			is_pointer = True
			break
	for x in valid_structure:
		if x in target:
			is_structure = True
			break
	# int * a
	# struct abc abc
	if ('struct' in target):
		struct_in_target = True
		target = target.replace('struct ', '')
	else:
		struct_in_target = False

	target_type = target[:target.find(' ')].replace(' ', '').replace(' ', '')
	if struct_in_target:
		target_type = 'struct {}'.format(target_type)
	else:
		pass

	target_val = target[target.find(' ') : ].replace(' ', '').replace('*', '').replace(';', '')
	return target_type, target_val, is_pointer, is_structure
#------------------------------------------------------------------------------#
def cell_extract_data(ws, cur_cell, target:str, cell_val:str, purpose:str)->str:
	cell_val = cell_val.replace(target, '')
	cell_type, cell_name, is_pointer, is_structure = get_type_name(ws, cell_val)
	print(cell_type)
	print(cell_name)
	print(is_pointer)
	print(is_structure)
	#--------------------------------------------------------------#
	data = ''
	if purpose == 'input_define':
		# define structure
		# struct CPPTH_LOOP_INPUT_STRUCT {
		#	 int a;
		#	 int * b;
		# }
		data = '\t\t{};\n'\
			.format(cell_val)
		pass
	#--------------------------------------------------------------#
	elif purpose == 'pointer_init_local':
		# define local variable - only for pointer
		# {cell_type} local_{cell_name};
		# uint32_t local_val;
		if is_pointer:
			data = '\t{} local_{};\n'\
			.format(cell_type, cell_name)
		pass
	#--------------------------------------------------------------#
	elif purpose == 'pointer_init':
		# define pointer init with local variable - only for pointer
		# if (CURRENT_TEST.a != NULL) {
		# 	 CURRENT_TEST.a = &local_a;
		# }
		if is_pointer:
			######
			data = '''\
			if (CURRENT_TEST.{pointer_name} != NULL) {{
				CURRENT_TEST.{pointer_name} = &local_{pointer_name};
			}}\n'''\
			.format(pointer_name = cell_name)
			######
		pass
	#--------------------------------------------------------------#
	elif purpose == 'pointer_check':
		# define check point of pointer output variable - only for pointer
		# only when [a]pointer is define in output range
		# CHECK_S_INT()
		check_address = ['IODevice']
		check_u_int = ['uint', 'unsigned']
		check_bool = ['bool', 'Boolean']
		# default check type is CHECK_S_INT
		check_type = 'CHECK_S_INT'
		for c in check_u_int:
			if c in cell_type:
				check_type = 'CHECK_U_INT'
		for c in check_bool:
			if c in cell_type:
				check_type = 'CHECK_BOOL'
		for c in check_address:
			if c in cell_type:
				check_type = 'CHECK_ADDRESS'
		if is_pointer == True and is_structure == False:
			######
			data = '''\
			{check_type}({left}, {right});\n'''\
			.format(check_type = check_type,\
					left = 'local_{}'.format(cell_name),\
					right = 'CURRENT_TEST.expected_{}'.format(cell_name))
			######
		elif is_structure == True:
			######
			if '[' in cell_name:
				cell_name = cell_name.replace('[', '_').replace(']', '')
			cell_upper = coor_shift_up(ws, cur_cell)
			cell_upper_val = get_cell_value(ws, cell_upper)
			cell_upper_type, cell_upper_name = get_type_name(ws, cell_upper_val)
			#print(cell_uper_val)
			data = '''\
			{check_type}({left}, {right});\n'''\
			.format(check_type = check_type,\
					left = '{}'.format(cell_upper_name),\
					right = 'CURRENT_TEST.expected_{}'.format(cell_name))
			######
			print(data)

	elif purpose == 'expected_define':
		if '[' in cell_name:
			cell_name = cell_name.replace('[', '_').replace(']', '')
		data = '''\
		{type} expected_{name};\n'''\
		.format(type = cell_type,\
			name = cell_name)
		pass
	#--------------------------------------------------------------#
	else:
		pass
	#--------------------------------------------------------------#
	return data
#------------------------------------------------------------------------------#
def structure_handle(ws, top_cell_name:str, target:str)->str:
	cell_1 = find_cell(ws, top_cell_name)
	cell_2 = coor_shift_down(ws, cell_1)
	data = ''
	while (cell_2['lastcol'] <= cell_1['lastcol']):
		cell_2_val = get_cell_value(ws, cell_2)
		if target in cell_2_val:
			cell_2_type, cell_2_name, is_pointer, is_structure = get_type_name(ws, cell_2_val.replace(target, ''))
			#------------------------------------------------------------------#
			# Check is this merged cell?
			if is_structure:
				if (cell_2['lastcol'] - cell_2['firstcol']) == 0:
					pass
				else:
					cell_3 = coor_shift_down(ws, cell_2)
					while (cell_3['lastcol'] <= cell_2['lastcol']):
						cell_3_val = get_cell_value(ws, cell_3)
						cell_3_type, cell_3_name, is_pointer, is_structure = get_type_name(ws, cell_3_val.replace(target, ''))
						if is_pointer:
							access = '->'
						else:
							access = '.'
						data_structre = '''\
			CURRENT_TEST.{name}{access}{element} = CURRENT_TEST.{element};\n'''\
						.format(\
							name = cell_2_name,\
							access = access,\
							element = cell_3_name\
						)
						data = '{}{}'.format(data, data_structre)
						cell_3 = coor_shift_right(ws, cell_3)
		cell_2 = coor_shift_right(ws, cell_2)
	return data
#------------------------------------------------------------------------------#
def search_argument(ws, top_cell_name:str, target:str, purpose:str)->str:
	cell_1 = find_cell(ws, top_cell_name)
	cell_2 = coor_shift_down(ws, cell_1)
	data = ''
	while (cell_2['lastcol'] <= cell_1['lastcol']):
		### Begin of while
		cell_2_val = get_cell_value(ws, cell_2)
		if target in cell_2_val:
			#------------------------------------------------------------------#
			# Check is this merged cell?
			if (cell_2['lastcol'] - cell_2['firstcol']) == 0:
				data = '{}{}'.format(data, cell_extract_data(ws, cell_2, target, cell_2_val, purpose))
				pass
			#------------------------------------------------------------------#
			# Handle merged cell
			# This will be structure
			else:
				cell_3 = coor_shift_down(ws, cell_2)
				while (cell_3['lastcol'] <= cell_2['lastcol']):
					cell_3_val = get_cell_value(ws, cell_3)
					data = '{}{}'.format(data, cell_extract_data(ws, cell_3, target, cell_3_val, purpose))
					###	Next of while
					cell_3 = coor_shift_right(ws, cell_3)
				pass
		###	Next of while
		cell_2 = coor_shift_right(ws, cell_2)
	return data
#------------------------------------------------------------------------------#
def get_input_argument(ws, top_cell_name:str, target:str)->str:
	cell_1 = find_cell(ws, top_cell_name)
	cell_2 = coor_shift_down(ws, cell_1)
	data = ''
	list_of_handle = []
	while (cell_2['lastcol'] <= cell_1['lastcol']):
		### Begin of while
		cell_2_val = get_cell_value(ws, cell_2)
		if target in cell_2_val:
			cell_type, cell_name, is_pointer, is_structure = get_type_name(ws, cell_2_val.replace(target, ''))
			if cell_name in list_of_handle:
				pass
			else:
				list_of_handle.append(cell_name)
				data_append = 'CURRENT_TEST.{},'.format(cell_name)
				data = '{}{}'.format(data, data_append)

		###	Next of while
		cell_2 = coor_shift_right(ws, cell_2)
	return data[:-1]
	pass
#------------------------------------------------------------------------------#
def test_program_format(ws, func_name:str):
	input_define = '{}{}{}{}'.format(\
		search_argument(ws, 'Input factor', '[a]', 'input_define')[:-1],\
		search_argument(ws, 'Input factor', '[g]', 'input_define')[:-1],\
		search_argument(ws, 'Output element', '[a]', 'expected_define')[:-1],\
		search_argument(ws, 'Output element', '[g]', 'expected_define')[:-1]\
	)
	print(input_define)
	pointer_init_local = '{}'.format(\
		search_argument(ws, 'Input factor', '[a]', 'pointer_init_local')[:-1]\
	)
	print(pointer_init_local)
	pointer_init = '{}{}'.format(\
		search_argument(ws, 'Input factor', '[a]', 'pointer_init'),\
		structure_handle(ws, 'Input factor', '[a]')
	)
	print(pointer_init)
	pointer_check = '{}{}'.format(\
		search_argument(ws, 'Output element', '[a]', 'pointer_check'),\
		search_argument(ws, 'Output element', '[g]', 'pointer_check')
	)
	print(pointer_check)
	input_argument = '{}'.format(\
		get_input_argument(ws, 'Input factor', '[a]')\
	)
	print(input_argument)
	data = '''\
void test_{func_name}(){{
	struct CPPTH_LOOP_INPUT_STRUCT {{
		/* Test case data declarations */
		char* name;
		char* description;
		char* expected_calls;
		int execute;
{input_define}
		int32_t expected_returnValue;
	}};
	int32_t returnValue;
	/* Import external data declarations */
	#include "test_{func_name}.h"

	/* Set global data */
	initialise_global_data();
	/* Set expected values for global data checks */
	initialise_expected_global_data();

{pointer_init_local}

	START_TEST_LOOP();
		/* Expected Call Sequence  */
		EXPECTED_CALLS(CURRENT_TEST.expected_calls);
{pointer_init}
			/* Call SUT */
			returnValue = {func_name}({input_argument});

			/* Test case checks */
			CHECK_S_INT(returnValue, CURRENT_TEST.expected_returnValue);
{pointer_check}
		END_CALLS();
	END_TEST_LOOP();
}}
'''.format(\
	func_name = func_name,\
	input_define = input_define,\
	pointer_init_local = pointer_init_local,\
	pointer_init = pointer_init,\
	pointer_check = pointer_check,\
	input_argument = input_argument\
	)
	return data

# Append text to file
def file_write(path, filename, data, position):
	write_done = False
	count = 0
	path_temp = '{}temp.txt'.format(path)
	path_file = '{}{}'.format(path, filename)
	try:
		os.rename(path_file, path_temp)
	except:
		sys.exit(0)
	with open(path_temp, 'r') as infile, open(path_file, 'w') as outfile:
		for line in infile:
			try:
				if (position[count] in line):
					count += 1
			except:
				pass
			if (count == len(position)) and (write_done == False):
				outfile.write(data)
				write_done = True
			outfile.write(line)
	os.remove(path_temp)

# Extract data from PCL with specific target
def get_data(ws, target, input_range, cur_row):
	data = ''
	input_cell = coor_shift_down(ws, input_range)
	while input_cell['lastcol'] <= input_range['lastcol']:
		# Check input cell merged range, if merge range use {var1, var2}
		if (target in get_cell_value(ws, input_cell)):
			if (input_cell['lastcol'] - input_cell['firstcol']) == 0:
				cur_input_param = get_cell_value(ws, [input_cell['firstcol'], cur_row])
				data = '{}{}, '.format(data, cur_input_param)
			else:
				'''
				data = '{}{{'.format(data)
				element_cell = coor_shift_down(ws, input_cell)
				while element_cell['lastcol'] <= input_cell['lastcol']:
					cur_input_param = get_cell_value(ws, [element_cell['firstcol'], cur_row])
					data = '{}{}, '.format(data, cur_input_param)
					element_cell = coor_shift_right(ws, element_cell)
				data = '{}}}, '.format(data[:-2])
				'''
				element_cell = coor_shift_down(ws, input_cell)
				while element_cell['lastcol'] <= input_cell['lastcol']:
					cur_input_param = get_cell_value(ws, [element_cell['firstcol'], cur_row])
					data = '{}{}, '.format(data, cur_input_param)
					element_cell = coor_shift_right(ws, element_cell)
		input_cell = coor_shift_right(ws, input_cell)
	return data

# Get test case row
def row_of_testcase(ws, symbol):
	cur_cell = find_cell(ws, symbol)
	col_of_tc = cur_cell['firstcol']
	first_row_of_tc = cur_cell['lastrow'] + 1
	while (get_cell_value(ws, cur_cell) != None):
		cur_cell = coor_shift_down(ws, cur_cell)
	last_row_of_tc = cur_cell['firstrow'] - 1
	return first_row_of_tc, last_row_of_tc, col_of_tc

# Create file .h contrain test case
def create_test_case_file(ws, worksheet, src_dir, src, check_sequence):
	# Get Test case start row and end row
	start_row, end_row, testcase_col = row_of_testcase(ws, '#')
	# Get Input factor range
	input_factor = find_cell(ws, 'Input factor')
	# Get Output element range
	output_element = find_cell(ws, 'Output element')
	# Create file .h

	dot_h_dir = '{}test_{}\\test_{}.h'.format(src_dir, src, worksheet)
	dot_h = open(dot_h_dir, 'w')

	# Begin of file
	data = 'static struct CPPTH_LOOP_INPUT_STRUCT CPPTH_LOOP_INPUT[] = {\n'

	dot_h.write(data)
	# Add test case
	# Input factor
	for cur_row in range(start_row, end_row + 1):
		# Reset data
		data = '\t{'
		# Add test case number to data
		tc_num = get_cell_value(ws, [testcase_col, cur_row])
		tc_num = tc_num[:tc_num.find('-')] + '_' + tc_num[tc_num.find('-') + 1:]
		data = '{}\"{}\", '.format(data, tc_num)
		# Add description to data - Named: Item
		describe = '{}_{}'.format(worksheet, tc_num)
		data = '{}\"{}\", '.format(data, describe)
		del describe

		# Add expected calls sequence
		# In case not check sequence of calling stub function
		# TODO: Add feature: many instance in sequence

		#data = data + '"'
		data = '{}"'.format(data)
		CELL = coor_shift_down(ws, input_factor)
		list_of_function_called = list()
		while CELL['lastcol'] <= input_factor['lastcol']:
			CELL_VAL = get_cell_value(ws, CELL)
			if ('[rt]' in CELL_VAL):
				# Get function name from [rt]Error MyFunction(int a, int b);
				#                                  MyFunction
				FNAME = CELL_VAL[CELL_VAL.find(' ') + 1 : CELL_VAL.find('(')]
				del CELL_VAL
				# Check loop of instance
				if (FNAME not in dict(list_of_function_called)):
					list_of_function_called.append([FNAME, 0])
				else:
					function_count = list_of_function_called[list(dict(list_of_function_called)).index(FNAME)][1]
					list_of_function_called[list(dict(list_of_function_called)).index(FNAME)][1] = function_count + 1
				# Check is this instance called
				FRETVAL = get_cell_value(ws, [CELL['firstcol'], cur_row])

				FUNCINSTANCE = ''
				if (FRETVAL != None) and (FRETVAL != '-'):
					FUNCINSTANCE = '{}#{}_{}'.format(FNAME, tc_num, list_of_function_called[list(dict(list_of_function_called)).index(FNAME)][1])

					if check_sequence == True:
						data = '{}{}; '.format(data, FUNCINSTANCE)
					else:
						data = '{}{{{}}} '.format(data, FUNCINSTANCE)

			CELL = coor_shift_right(ws, CELL)
		data = '{}\", '.format(data)

		# Add execute - 1: execute this function
		judgment = find_cell(ws, 'Judgment')
		if ('Exclude' in get_cell_value(ws, [judgment['firstcol'], cur_row])):
			data = '{}0, '.format(data)
		else:
			data = '{}1, '.format(data)

		# Add input param by detect [a]
		# TODO: Add detect structure - DONE

		data = data + get_data(ws, '[a]', input_factor, cur_row)

		# Add global variable by detect [g]
		data = data + get_data(ws, '[g]', input_factor, cur_row)

		# TODO: Add detect check [a] param output - DONE
		data = data + get_data(ws, '[a]', output_element, cur_row)

		# Add expected global variable by detect [g] in output element
		data = data + get_data(ws, '[g]', output_element, cur_row)

		# Add test result by detect 'Return value' in output element
		data = data + get_data(ws, 'Return value', output_element, cur_row)

		# Write all the data to file
		#data = data[:-2] + '},\n'
		data = '{}}},\n'.format(data[:-2])
		dot_h.write(data)

	# End of file
	data = '};\n'
	dot_h.write(data)
	dot_h.close()
	del dot_h

# Create stub instance
def create_stub_file(ws, worksheet, src_dir, src):
	# Create stub function
	#dot_c = open('test_' + worksheet + '.c', 'w')
	# Get the first test case's row, and the last test case's row, and the current col
	start_row, end_row, testcase_col = row_of_testcase(ws, '#')
	# Get Input factor range
	input_factor = find_cell(ws, 'Input factor')
	# Get Output element range
	output_element = find_cell(ws, 'Output element')

	for cur_row in range(start_row, end_row + 1):
		# Get test case number
		tc_num = get_cell_value(ws, [testcase_col, cur_row])
		tc_num = tc_num[:tc_num.find('-')] + '_' + tc_num[tc_num.find('-') + 1:]
		# Create instance for test case num

		list_of_function_called = list()
		input_cell = coor_shift_down(ws, input_factor)
		while input_cell['lastcol'] <= input_factor['lastcol']:
			# Check [rt]
			#cur_input_param = get_cell_value(ws, [input_cell['firstcol'], cur_row])
			CELL_VAL = get_cell_value(ws, input_cell)
			if ('[rt]' in CELL_VAL):
				FNAME = CELL_VAL[CELL_VAL.find(' ') + 1 : CELL_VAL.find('(')]
				del CELL_VAL
				# Check loop of instance
				if (FNAME not in dict(list_of_function_called)):
					list_of_function_called.append([FNAME, 0])
				else:
					function_count = list_of_function_called[list(dict(list_of_function_called)).index(FNAME)][1]
					list_of_function_called[list(dict(list_of_function_called)).index(FNAME)][1] = function_count + 1

				void_return_type = False

				# Get function name
				FNAME = get_cell_value(ws, input_cell)

				if 'void' in FNAME[:FNAME.find('(')]:
					void_return_type = True

				# Extract function name from [rt]type function_name(...);
				FNAME = FNAME[FNAME.find(' ') + 1: FNAME.find('(')]

				# TODO: title is Isolate, need to implement for other title, Stub, Wrapper
				title = '/* Isolate for function %s */\n' %(FNAME)
				# Set instance
				if_instance = '\tIF_INSTANCE(\"%s\") {\n' %(tc_num + '_' + str(list_of_function_called[list(dict(list_of_function_called)).index(FNAME)][1]))
				outval_data = ''

				# CHECK DATA input of stub function
				check_data = ''
				output_cell = coor_shift_down(ws, output_element) 
				while output_cell['lastcol'] <= output_element['lastcol']:
					# Check [f] symbol and check the function name
					if ('[f]' in get_cell_value(ws, output_cell)) and (FNAME in get_cell_value(ws, output_cell)):
						check_point = coor_shift_down(ws, output_cell)
						while check_point['lastcol'] <= output_cell['lastcol']:

							# TODO: add replace 'CHECK_S_INT' - Done
							check_point_val = str(get_cell_value(ws, [check_point['firstcol'] , cur_row]))
							if (check_point_val != None) and (check_point_val) != '-':

								# Classify the check value
								if ('UTS_NON0' in check_point_val) or ('false' in check_point_val) or ('true' in check_point_val): 
									#check_data = check_data + '\t\t' + 'CHECK_BOOLEAN' + '((' + str(get_cell_value(ws, check_point)) + ' != NULL)' + ', true);\n'
									check_data = '{}\t\tCHECK_BOOLEAN(({} != NULL), true);\n'.format(check_data, get_cell_value(ws, check_point))
								elif ('NULL' in check_point_val) or (check_point_val.isupper()) or ('&' in check_point_val) or ('Connection' in check_point_val):
									#check_data = check_data + '\t\t' + 'CHECK_ADDRESS' + '(' + str(get_cell_value(ws, check_point)) + ', ' + check_point_val + ');\n'
									check_data = '{}\t\tCHECK_ADDRESS({}, {});\n'.format(check_data, get_cell_value(ws, check_point), check_point_val)
								elif (check_point_val.isdigit()):
									#check_data = check_data + '\t\t' + 'CHECK_S_INT' + '(' + str(get_cell_value(ws, check_point) )+ ', ' + check_point_val + ');\n'
									check_data = '{}\t\tCHECK_S_INT({}, {});\n'.format(check_data, get_cell_value(ws, check_point), check_point_val)
								elif ('u' or 'U' in check_point_val):
									#check_data = check_data + '\t\t' + 'CHECK_U_INT' + '(' + str(get_cell_value(ws, check_point)) + ', ' + check_point_val + ');\n'
									check_data = '{}\t\tCHECK_U_INT({}, {});\n'.format(check_data, get_cell_value(ws, check_point), check_point_val)
								else:
									#check_data = check_data + '\t\t' + 'CHECK_ADDRESS' + '(' + str(get_cell_value(ws, check_point)) + ', ' + check_point_val + ');\n'
									check_data = '{}\t\tCHECK_ADDRESS({}, {});\n'.format(check_data, get_cell_value(ws, check_point), check_point_val)

							check_point = coor_shift_right(ws, check_point)
					output_cell = coor_shift_right(ws, output_cell)

				# Function output value
				if input_cell['lastcol'] < input_factor['lastcol']:
					outval_range = coor_shift_right(ws, input_cell)
					if ('[f]' in get_cell_value(ws, outval_range)):
						outval_cell = coor_shift_down(ws, outval_range)
						while outval_cell['lastcol'] <= outval_range['lastcol']:
							outval = str(get_cell_value(ws, [outval_cell['firstcol'], cur_row]))
							if (outval != None) and (outval != '-'):
								if ('UTS' in outval):
									return_val = get_cell_value(ws, outval_cell)
									#outval_data = outval_data + '\t\t' + '*' + str(get_cell_value(ws, outval_cell)) + ' = ' \
									#+ outval[:outval.find(')') + 1] + '&local' + ';\n'
									if '*' in return_val:
										outval_data = '{}\t\t{} = {}&local;\n'.format(outval_data, return_val, outval[:outval.find(')') + 1])
									else:
										outval_data = '{}\t\t*{} = {}&local;\n'.format(outval_data, return_val, outval[:outval.find(')') + 1])
								else:
									#outval_data = outval_data + '\t\t' + '*' + str(get_cell_value(ws, outval_cell)) + ' = ' + outval + ';\n'
									return_val = get_cell_value(ws, outval_cell)
									if '*' in return_val:
										outval_data = '{}\t\t{} = {};\n'.format(outval_data, return_val, outval)
									else:
										outval_data = '{}\t\t*{} = {};\n'.format(outval_data, return_val, outval)
							outval_cell = coor_shift_right(ws, outval_cell)

				# Set return value for Stub function
				return_value = get_cell_value(ws, [input_cell['firstcol'], cur_row])

				instance_existed = True

				if void_return_type == True:
					data_return = '\t\treturn;\n\t}\n'
				else:
					if return_value != None and return_value != '-':
						#data_return = '\t\treturn ' + str(return_value) + ';\n\t}\n'
						data_return = '\t\treturn {};\n\t}}\n'.format(return_value)
					else:
						# This instance is not used
						instance_existed = False
						#data_return = '\t\treturn ' + '0' + ';\n\t}\n'

				if instance_existed:
					data = if_instance + check_data + outval_data + data_return

					# Get position for append data to source
					position = [title, 'IF_INSTANCE("default")', '}', 'LOG_SCRIPT_ERROR']
					# Write file to test program of Cantata
					#file_write(src_dir, 'test_' + src + '.c', data, position)
					file_write(src_dir, 'test_{}.c'.format(src) , data, position)

					data = title + data
					#dot_c.write(data)

			input_cell = coor_shift_right(ws, input_cell)
	#dot_c.close()

# Main function
def main(argv):
	try:
		opts, args = getopt.getopt(argv,"hid:i:ws:sr:se:ca:",["idir=","ifile=","wsheet=","source=","check_seq=","can_dir="])
	except getopt.GetoptError:
		print("--idir <inputfir> --ifile <inputfile> --wsheet <worksheet> --check_seq <True/False> --source <source>")
		sys.exit(2)
	for opt, arg in opts:
		if opt == '-h':
			print("--idir <inputfir> --ifile <inputfile> --wsheet <worksheet>")
			sys.exit()
		elif opt in ("-id", "--idir"):
			inputdir = arg
		elif opt in ("-i", "--ifile"):
			inputfile = arg
		elif opt in ("-ws", "--wsheet"):
			worksheet = arg
		elif opt in ("-sr", "--source"):
			source = arg
		elif opt in ("-se", "--check_seq"):
			if ('rue' in arg):
				check_sequence = True
			else:
				check_sequence = False
		elif opt in ("-ca", "--can_dir"):
			src_dir = arg + '\\'
	print(inputdir)
	print(inputfile)
	# Get working sheet
	ws = load_worksheet('{}\\{}'.format(inputdir, inputfile), worksheet)

	# Create file dot h, contain all the test case
	create_test_case_file(ws, worksheet, src_dir, source, check_sequence)

	# Create stub function
	src_dir = '{}test_{}\\'.format(src_dir, source)
	create_stub_file(ws, worksheet, src_dir, source)

	# Create test program
	data = test_program_format(ws, worksheet)
	position = ['/* Call Interface Control                                                    */',\
		'/*****************************************************************************/']
	#file_write(src_dir, 'test_{}.c'.format(source) , data, position)
	#print(data)

if __name__ == "__main__":
	main(sys.argv[1:])
