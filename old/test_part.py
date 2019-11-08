import os, sys, getopt
from openpyxl import load_workbook

#------------------------------------------------------------------------------#
# Find cell boundary
def find_cell(ws, target, match_case=False, find_all=False):
	return_value = list()
	target_found = False
	# If target is string type
	if isinstance(target, str):
		for rows in ws:
			for cell in rows:
				if match_case:
					target_found = target in cell.value
				else:
					target_found = target == cell.value

				if target_found:
					not_merged_cell = True
					for merged_cell in ws.merged_cells:
						if (cell.coordinate in merged_cell):
							temp = {'firstcol':merged_cell.bounds[0],
									'firstrow':merged_cell.bounds[1],
									'lastcol':merged_cell.bounds[2],
									'lastrow':merged_cell.bounds[3]}
							not_merged_cell = False
							if find_all:
								return_value.append(temp)
							else:
								return dict(temp)
					if not_merged_cell:
						temp = {'firstcol':cell.column,
								'firstrow':cell.row,
								'lastcol':cell.column,
								'lastrow':cell.row}
						if find_all:
							return_value.append(temp)
						else:
							return dict(temp)
	# If target is list type, list format [column, row]
	elif isinstance(target, list):
		# list  = [col, row]
		coor = ws.cell(column=target[0], row=target[1]).coordinate
		not_merged_cell = True
		for merged_cell in ws.merged_cells:
			if (coor in merged_cell):
				temp = {'firstcol':merged_cell.bounds[0],
						'firstrow':merged_cell.bounds[1],
						'lastcol':merged_cell.bounds[2],
						'lastrow':merged_cell.bounds[3]}
				not_merged_cell = False
				return dict(temp)
		if not_merged_cell:
			temp = {'firstcol':target[0],
					'firstrow':target[1],
					'lastcol':target[0],
					'lastrow':target[1]}
			return dict(temp)

	return return_value
#------------------------------------------------------------------------------#
# Get cell value with cell coordinate
def get_cell_value(ws, coor)->str:
	if isinstance(coor, list):
		val = ws.cell(column=coor[0], row=coor[1]).value
		return val
	else:
		for row in range (coor['firstrow'], coor['lastrow'] + 1):
			for col in range (coor['firstcol'], coor['lastcol'] + 1):
				val = ws.cell(column=col, row=row).value
				if (val != None):
					return val
	return None
#------------------------------------------------------------------------------#
# Shift 'right' cell
def coor_shift_right(ws, coor):
	temp = list([coor['lastcol'] + 1, coor['firstrow']])
	return find_cell(ws, temp)
#------------------------------------------------------------------------------#
# Shift 'left' cell
def coor_shift_left(ws, coor):
	col = coor['firstcol'] - 1
	if col < 1:
		col = 1
	temp = list([col, coor['firstrow']])
	return find_cell(ws, temp)
#------------------------------------------------------------------------------#
# Shift 'up' cell
def coor_shift_up(ws, coor):
	row = coor['firstrow'] - 1
	if row < 1:
		row = 1
	temp = list([coor['firstcol'], row])
	return find_cell(ws, temp)
#------------------------------------------------------------------------------#
# Shift 'down' cell
def coor_shift_down(ws, coor):
	temp = list([coor['firstcol'], coor['lastrow'] + 1])
	return find_cell(ws, temp)
#------------------------------------------------------------------------------#
def load_worksheet(filename, sheetname):
	try:
		wb = load_workbook(filename, data_only=True)
	except:
		sys.exit(0)
	ws = wb[sheetname]
	return ws
#------------------------------------------------------------------------------#
def get_type_name(ws, target:str):
	valid_pointer = ['*', 'IODevice']
	valid_structure = ['struct', 'g_struct']
	is_pointer = False
	is_structure = False
	for x in valid_pointer:
		if x in target:
			is_pointer = True
			break
	for x in valid_structure:
		if x in target:
			is_structure = True
			break
	# int * a
	# struct abc abc
	if ('struct' in target):
		struct_in_target = True
		target = target.replace('struct ', '')
	else:
		struct_in_target = False
	
	target_type = target[:target.find(' ')].replace(' ', '').replace(' ', '')
	if struct_in_target:
		target_type = 'struct {}'.format(target_type)
	else:
		pass

	target_val = target[target.find(' ') : ].replace(' ', '').replace('*', '').replace(';', '')
	return target_type, target_val, is_pointer, is_structure
#------------------------------------------------------------------------------#
def cell_extract_data(target:str, cell_val:str, purpose:str)->str:
	cell_val = cell_val.replace(target, '')
	cell_type, cell_name, is_pointer, is_structure = get_type_name(ws, cell_val)
	#--------------------------------------------------------------#
	data = ''
	if purpose == 'input_define':
		# define structure
		# struct CPPTH_LOOP_INPUT_STRUCT {
		#	 int a;
		#	 int * b;
		# }
		data = '\t\t{};\n'\
			.format(cell_val)
		pass
	#--------------------------------------------------------------#
	elif purpose == 'pointer_init_local':
		# define local variable - only for pointer
		# {cell_type} local_{cell_name};
		# uint32_t local_val;
		if is_pointer:
			data = '\t{} local_{};\n'\
			.format(cell_type, cell_name)
		pass
	#--------------------------------------------------------------#
	elif purpose == 'pointer_init':
		# define pointer init with local variable - only for pointer
		# if (CURRENT_TEST.a != NULL) {
		# 	 CURRENT_TEST.a = &local_a;
		# }
		if is_pointer:
			######
			data = '''\
			if (CURRENT_TEST.{pointer_name} != NULL) {{
				CURRENT_TEST.{pointer_name} = &local_{pointer_name};
			}}\n'''\
			.format(pointer_name = cell_name)
			######
		pass
	#--------------------------------------------------------------#
	elif purpose == 'pointer_check':
		# define check point of pointer output variable - only for pointer
		# only when [a]pointer is define in output range
		# CHECK_S_INT()
		check_address = ['IODevice']
		check_u_int = ['uint', 'unsigned']
		check_bool = ['bool', 'Boolean']
		if is_pointer:
			# default check type is CHECK_S_INT
			check_type = 'CHECK_S_INT'
			for c in check_u_int:
				if c in cell_type:
					check_type = 'CHECK_U_INT'
			for c in check_bool:
				if c in cell_type:
					check_type = 'CHECK_BOOL'
			for c in check_address:
				if c in cell_type:
					check_type = 'CHECK_ADDRESS'
			######
			data = '''\
			{check_type}({left}, {right});\n'''\
			.format(check_type = check_type,\
					left = 'local_{}'.format(cell_name),\
					right = 'CURRENT_TEST.expected_{}'.format(cell_name))
			######
		pass
	elif purpose == 'expected_define':
		data = '''\
		{type} expected_{name};\n'''\
		.format(type = cell_type,\
			name = cell_name)
		pass
	#--------------------------------------------------------------#
	else:
		pass
	#--------------------------------------------------------------#
	return data
#------------------------------------------------------------------------------#
def structure_handle(ws, top_cell_name:str, target:str)->str:
	cell_1 = find_cell(ws, top_cell_name)
	cell_2 = coor_shift_down(ws, cell_1)
	data = ''
	while (cell_2['lastcol'] <= cell_1['lastcol']):
		cell_2_val = get_cell_value(ws, cell_2)
		if target in cell_2_val:
			cell_2_type, cell_2_name, is_pointer, is_structure = get_type_name(ws, cell_2_val.replace(target, ''))
			#------------------------------------------------------------------#
			# Check is this merged cell?
			if is_structure:
				if (cell_2['lastcol'] - cell_2['firstcol']) == 0:
					pass
				else:
					cell_3 = coor_shift_down(ws, cell_2)
					while (cell_3['lastcol'] <= cell_2['lastcol']):
						cell_3_val = get_cell_value(ws, cell_3)
						cell_3_type, cell_3_name, is_pointer, is_structure = get_type_name(ws, cell_3_val.replace(target, ''))
						if is_pointer:
							access = '->'
						else:
							access = '.'
						data_structre = '''\
			CURRENT_TEST.{name}{access}{element} = CURRENT_TEST.{element};\n'''\
						.format(\
							name = cell_2_name,\
							access = access,\
							element = cell_3_name\
						)
						data = '{}{}'.format(data, data_structre)
						cell_3 = coor_shift_right(ws, cell_3)
		cell_2 = coor_shift_right(ws, cell_2)
	return data
#------------------------------------------------------------------------------#
def search_argument(ws, top_cell_name:str, target:str, purpose:str)->str:
	cell_1 = find_cell(ws, top_cell_name)
	cell_2 = coor_shift_down(ws, cell_1)
	data = ''
	while (cell_2['lastcol'] <= cell_1['lastcol']):
		### Begin of while
		cell_2_val = get_cell_value(ws, cell_2)
		if target in cell_2_val:
			#------------------------------------------------------------------#
			# Check is this merged cell?
			if (cell_2['lastcol'] - cell_2['firstcol']) == 0:
				data = '{}{}'.format(data, cell_extract_data(target, cell_2_val, purpose))
				pass
			#------------------------------------------------------------------#
			# Handle merged cell
			# This will be structure
			else:
				cell_3 = coor_shift_down(ws, cell_2)
				while (cell_3['lastcol'] <= cell_2['lastcol']):
					cell_3_val = get_cell_value(ws, cell_3)
					data = '{}{}'.format(data, cell_extract_data(target, cell_3_val, purpose))
					###	Next of while
					cell_3 = coor_shift_right(ws, cell_3)
				pass
		###	Next of while
		cell_2 = coor_shift_right(ws, cell_2)
	return data
#------------------------------------------------------------------------------#
def get_input_argument(ws, top_cell_name:str, target:str)->str:
	valid_top_cell_name = ['Input factor', 'Output element']
	valid_target = ['[a]', '[f]', '[rt]', '[g]']
	if (target not in valid_target) or (top_cell_name not in valid_top_cell_name):
		return
	del valid_top_cell_name, valid_target

	cell_1 = find_cell(ws, top_cell_name)
	cell_2 = coor_shift_down(ws, cell_1)
	data = ''
	list_of_handle = []
	while (cell_2['lastcol'] <= cell_1['lastcol']):
		### Begin of while
		cell_2_val = get_cell_value(ws, cell_2)
		if target in cell_2_val:
			cell_type, cell_name, is_pointer, is_structure = get_type_name(ws, cell_2_val.replace(target, ''))
			if cell_name in list_of_handle:
				pass
			else:
				list_of_handle.append(cell_name)
				data_append = 'CURRENT_TEST.{},'.format(cell_name)
				data = '{}{}'.format(data, data_append)

		###	Next of while
		cell_2 = coor_shift_right(ws, cell_2)
	return data[:-1]
	pass
#------------------------------------------------------------------------------#
def test_program_format(ws, func_name:str):
	input_define = '{}{}'.format(\
		search_argument(ws, 'Input factor', '[a]', 'input_define')[:-1],\
		search_argument(ws, 'Output element', '[a]', 'expected_define')[:-1],\
	)
	pointer_init_local = '{}'.format(\
		search_argument(ws, 'Input factor', '[a]', 'pointer_init_local')[:-1]\
	)
	pointer_init = '{}{}'.format(\
		search_argument(ws, 'Input factor', '[a]', 'pointer_init'),\
		structure_handle(ws, 'Input factor', '[a]')
	)
	pointer_check = '{}{}'.format(\
		search_argument(ws, 'Output element', '[a]', 'pointer_check'),\
		search_argument(ws, 'Output element', '[g]', 'pointer_check')
	)
	input_argument = '{}'.format(\
		get_input_argument(ws, 'Input factor', '[a]')\
	)
	data = '''\
void test_{func_name}(){{
	struct CPPTH_LOOP_INPUT_STRUCT {{
		/* Test case data declarations */
		char* name;
		char* description;
		char* expected_calls;
		int execute;
{input_define}
		int32_t expected_returnValue;
	}};
	int32_t returnValue;
	/* Import external data declarations */
	#include "test_{func_name}.h"

	/* Set global data */
	initialise_global_data();
	/* Set expected values for global data checks */
	initialise_expected_global_data();

{pointer_init_local}

	START_TEST_LOOP();
		/* Expected Call Sequence  */
		EXPECTED_CALLS(CURRENT_TEST.expected_calls);
{pointer_init}
			/* Call SUT */
			returnValue = {func_name}({input_argument});

			/* Test case checks */
			CHECK_S_INT(returnValue, CURRENT_TEST.expected_returnValue);
{pointer_check}
		END_CALLS();
	END_TEST_LOOP();
}}
'''.format(\
	func_name = func_name,\
	input_define = input_define,\
	pointer_init_local = pointer_init_local,\
	pointer_init = pointer_init,\
	pointer_check = pointer_check,\
	input_argument = input_argument\
	)
	with open('texting.c', 'w') as f:
		f.write(data)
#------------------------------------------------------------------------------#
# Testing function
#ws = load_worksheet('test.xlsx', 'Sheet1')
#test_program_format(ws, 'r_imr_osal_init')