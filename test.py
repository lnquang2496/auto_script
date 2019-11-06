import os, sys, getopt
from openpyxl import load_workbook

# Find cell boundary
def find_cell(ws, target, match_case=False, find_all=False):
	return_value = list()
	target_found = False
	# If target is string type
	if isinstance(target, str):
		for rows in ws:
			for cell in rows:
				if match_case:
					target_found = target in cell.value
				else:
					target_found = target == cell.value

				if target_found:
					not_merged_cell = True
					for merged_cell in ws.merged_cells:
						if (cell.coordinate in merged_cell):
							temp = {'firstcol':merged_cell.bounds[0],
									'firstrow':merged_cell.bounds[1],
									'lastcol':merged_cell.bounds[2],
									'lastrow':merged_cell.bounds[3]}
							not_merged_cell = False
							if find_all:
								return_value.append(temp)
							else:
								return dict(temp)
					if not_merged_cell:
						temp = {'firstcol':cell.column,
								'firstrow':cell.row,
								'lastcol':cell.column,
								'lastrow':cell.row}
						if find_all:
							return_value.append(temp)
						else:
							return dict(temp)
	# If target is list type, list format [column, row]
	elif isinstance(target, list):
		# list  = [col, row]
		coor = ws.cell(column=target[0], row=target[1]).coordinate
		not_merged_cell = True
		for merged_cell in ws.merged_cells:
			if (coor in merged_cell):
				temp = {'firstcol':merged_cell.bounds[0],
						'firstrow':merged_cell.bounds[1],
						'lastcol':merged_cell.bounds[2],
						'lastrow':merged_cell.bounds[3]}
				not_merged_cell = False
				return dict(temp)
		if not_merged_cell:
			temp = {'firstcol':target[0],
					'firstrow':target[1],
					'lastcol':target[0],
					'lastrow':target[1]}
			return dict(temp)

	return return_value

# Get cell value with cell coordinate
def get_cell_value(ws, coor):
	if isinstance(coor, list):
		val = ws.cell(column=coor[0], row=coor[1]).value
		return str(val)
	else:
		for row in range (coor['firstrow'], coor['lastrow'] + 1):
			for col in range (coor['firstcol'], coor['lastcol'] + 1):
				val = ws.cell(column=col, row=row).value
				if (val != None):
					return str(val)
	return None

# Shift 'right' cell
def coor_shift_right(ws, coor):
	temp = list([coor['lastcol'] + 1, coor['firstrow']])
	return find_cell(ws, temp)

# Shift 'left' cell
def coor_shift_left(ws, coor):
	col = coor['firstcol'] - 1
	if col < 1:
		col = 1
	temp = list([col, coor['firstrow']])
	return find_cell(ws, temp)

# Shift 'up' cell
def coor_shift_up(ws, coor):
	row = coor['firstrow'] - 1
	if row < 1:
		row = 1
	temp = list([coor['firstcol'], row])
	return find_cell(ws, temp)

# Shift 'down' cell
def coor_shift_down(ws, coor):
	temp = list([coor['firstcol'], coor['lastrow'] + 1])
	return find_cell(ws, temp)

def load_worksheet(filename, sheetname):
	try:
		wb = load_workbook(filename, data_only=True)
	except:
		sys.exit(0)
	ws = wb[sheetname]
	return ws

def main(argv):
	try:
		opts, args = getopt.getopt(argv,"hid:i:ws:sr:se:ca:",["idir=","ifile=","wsheet=","source=","check_seq=","can_dir="])
	except getopt.GetoptError:
		print("--idir <inputfir> --ifile <inputfile> --wsheet <worksheet> --check_seq <True/False> --source <source>")
		sys.exit(2)
	for opt, arg in opts:
		if opt == '-h':
			print("--idir <inputfir> --ifile <inputfile> --wsheet <worksheet>")
			sys.exit()
		elif opt in ("-id", "--idir"):
			inputdir = arg
		elif opt in ("-i", "--ifile"):
			inputfile = arg
		elif opt in ("-ws", "--wsheet"):
			worksheet = arg
		elif opt in ("-sr", "--source"):
			source = arg
		elif opt in ("-se", "--check_seq"):
			if ('rue' in arg):
				check_sequence = True
			else:
				check_sequence = False
		elif opt in ("-ca", "--can_dir"):
			src_dir = arg + '\\'

	ws = load_worksheet('{}\\{}'.format(inputdir, inputfile), worksheet)

	input_param = ''
	pointer_init = ''
	pointer_init_local = ''

	input_factor = find_cell(ws, 'Input factor')
	# Get Output element range
	output_element = find_cell(ws, 'Output element')
#########################################################
	CELL = coor_shift_down(ws, input_factor)
	while CELL['lastcol'] <= input_factor['lastcol']:
		CELL_VAL = get_cell_value(ws, CELL)
		
		if ('[a]' in CELL_VAL):
			if (CELL['lastcol'] - CELL['firstcol']) == 0:

				if ('*' in CELL_VAL):
					temp_val = CELL_VAL

					type_of_val = temp_val[temp_val.find(']') + 1 : temp_val.find('*')]
					name_of_val = temp_val[temp_val.find('*') + 1 : ].replace(' ', '')

					pointer_init_local_temp = \
'''
\t{type} local_{name};'''\
.format(type= type_of_val, name=name_of_val)

					pointer_init_local = '{}{}'.format(pointer_init_local, pointer_init_local_temp)
					pointer_init_temp = \
'''
\t\t\tif (CURRENT_TEST.{name} != NULL) {{
\t\t\t\tCURRENT_TEST.{name} = &local_{name};
\t\t\t}}'''\
.format(name= name_of_val)

					pointer_init = '{}{}'.format(pointer_init, pointer_init_temp)
					pass
				else:
					pass

				CELL_VAL = CELL_VAL[CELL_VAL.find(']') + 1 : ]
				if CELL_VAL in input_param:
					pass
				else:
					input_param = '{}\t\t{};\n'.format(input_param, CELL_VAL)

			else:
				struct_element = coor_shift_down(ws, CELL)
				while struct_element['lastcol'] <= CELL['lastcol']:
					struct_element_cell = get_cell_value(ws, struct_element)
					

					input_param = '{}\t\t{};\n'.format(input_param, struct_element_cell)

					struct_name = get_cell_value(ws, CELL)
					struct_name = struct_name[struct_name.find('*') + 1 : ].replace(' ', '')
					element = struct_element_cell[struct_element_cell.find(' ') : ].replace(' ', '').replace('*', '')
					

					# TODO: Check struct is pointer or not.
					pointer_init_data = \
'''
\t\t\tCURRENT_TEST.{struct_name}->{element} = CURRENT_TEST.{element};'''\
.format(struct_name=struct_name, element=element)
					pointer_init = '{}{}'.format(pointer_init, pointer_init_data)

					struct_element = coor_shift_right(ws, struct_element)
				pass

		CELL = coor_shift_right(ws, CELL)
############################################
	pointer_check = ''
	CELL = coor_shift_down(ws, output_element)
	while CELL['lastcol'] <= output_element['lastcol']:
		cell_val = get_cell_value(ws, CELL)
		if ('[a]' in cell_val):
			cell_val = cell_val.replace('[a]', '')
			print(cell_val)
			name_of_val = cell_val[cell_val.find(' '):].replace(' ', '').replace('*', '')
			left = 'local_{}'.format(name_of_val)
			right = 'CURRENT_TEST.{}'.format(name_of_val)
			check_data = \
'''
\t\t\tCHECK_S_INT({left}, {right});'''\
.format(left = left, right = right)

			pointer_check = '{}{}'.format(pointer_check, check_data)
		CELL = coor_shift_right(ws, CELL)

#############################################
	data =\
'''void test_{func_name}(){{
	struct CPPTH_LOOP_INPUT_STRUCT {{
		/* Test case data declarations */
		char* name;
		char* description;
		char* expected_calls;
{input_param}
		int32_t expected_returnValue;
	}};
	int32_t returnValue;
	/* Import external data declarations */
	#include "test_{func_name}.h"

	/* Set global data */
	initialise_global_data();
	/* Set expected values for global data checks */
	initialise_expected_global_data();

{pointer_init_local}

	START_TEST_LOOP();
		/* Expected Call Sequence  */
		EXPECTED_CALLS(CURRENT_TEST.expected_calls);
{pointer_init}
			/* Call SUT */
			returnValue = {func_name}();

			/* Test case checks */
			CHECK_S_INT(returnValue, CURRENT_TEST.expected_returnValue);
			{pointer_check}
		END_CALLS();
	END_TEST_LOOP();
}}

'''.format(func_name=worksheet, input_param=input_param, pointer_init=pointer_init, pointer_check=pointer_check, pointer_init_local=pointer_init_local)
	with open('temp.c', 'w') as f:
		f.write(data)

if __name__ == "__main__":
	main(sys.argv[1:])